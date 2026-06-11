"""Tests for reporting/metrics.py with a 40-trade synthetic fixture.

Fixture layout (10 trades per strategy × 4 strategies = 40 total):
  4 × TP2  pnl=+18.0 pips   mae=3.0  mfe=20.0
  2 × BE   pnl=+0.5  pips   mae=2.0  mfe=10.0
  3 × SL   pnl=-15.0 pips   mae=15.0 mfe=5.0
  1 × TIMEOUT pnl=-3.0 pips mae=5.0  mfe=8.0
  → 6 wins / 4 losses per strategy → winrate=0.6

Expected metrics per strategy:
  avg_win  = (4×18 + 2×0.5) / 6 = 73/6 ≈ 12.167
  avg_loss = (3×(-15) + 1×(-3)) / 4 = -48/4 = -12.0
  expectancy = 0.6×12.167 + 0.4×(-12) = 7.3 - 4.8 = 2.5
  profit_factor = 73 / 48 ≈ 1.521
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import uuid
from datetime import datetime, timezone, timedelta

import pytest

import core.store as store
from core.models import Signal, TradeRecord
from reporting.metrics import (
    compute_stats, compute_funnel, compute_by_killzone, compute_daily,
    TradeStats, SignalFunnel, DayRow,
)


# ── Fixture ───────────────────────────────────────────────────────────────────

@pytest.fixture(autouse=True)
def _tmp_db(tmp_path, monkeypatch):
    monkeypatch.setattr(store, "DB_PATH", tmp_path / "test.db")
    store.init_db()


# trade templates per strategy
_TRADE_TEMPLATES = [
    # (exit_reason, pnl_pips, mae_pips, mfe_pips, duration_min, count)
    ("TP2",     18.0,  3.0, 20.0, 28, 4),
    ("BE",       0.5,  2.0, 10.0, 22, 2),
    ("SL",     -15.0, 15.0,  5.0, 42, 3),
    ("TIMEOUT", -3.0,  5.0,  8.0, 46, 1),
]

# strategy → (killzone, direction)
_STRAT_META = {
    "S1": ("LONDON",  "LONG"),
    "S2": ("NY_AM",   "LONG"),
    "S3": ("ASIA",    "LONG"),
    "S4": ("LONDON",  "SHORT"),
}


@pytest.fixture()
def seeded(tmp_path, monkeypatch):
    """Insert 40 trades + signals + funnel-relevant skipped signals."""
    monkeypatch.setattr(store, "DB_PATH", tmp_path / "seed.db")
    store.init_db()

    base_entry = datetime(2024, 3, 4, 10, 0, tzinfo=timezone.utc)  # Monday

    for strategy, (killzone, direction) in _STRAT_META.items():
        trade_num = 0
        for exit_reason, pnl, mae, mfe, dur, count in _TRADE_TEMPLATES:
            for _ in range(count):
                sig = Signal(
                    signal_id=uuid.uuid4().hex,
                    ts_utc=base_entry + timedelta(hours=trade_num),
                    strategy=strategy, direction=direction, killzone=killzone,
                    entry_type="MARKET", entry_price=2000.0,
                    entry_zone_low=1999.5, entry_zone_high=2000.5,
                    sl=1998.0 if direction == "LONG" else 2002.0,
                    tp1=2001.0 if direction == "LONG" else 1999.0,
                    tp2=2002.0 if direction == "LONG" else 1998.0,
                    sl_pips=20.0, confluences=[], score=3, context={},
                )
                store.insert_signal(sig, status="EXECUTED")

                entry_ts = base_entry + timedelta(hours=trade_num)
                exit_ts  = entry_ts + timedelta(minutes=dur)
                entry_price = 2000.0
                exit_price  = entry_price + (pnl * 0.10 if direction == "LONG"
                                             else -pnl * 0.10)

                trade = TradeRecord(
                    trade_id=uuid.uuid4().hex,
                    signal_id=sig.signal_id,
                    mt5_ticket=10000 + trade_num,
                    strategy=strategy, direction=direction,
                    lot=0.2, entry_price_fill=entry_price,
                    entry_ts_utc=entry_ts, sl_initial=1998.0, sl_current=1998.0,
                    tp1=2001.0, tp2=2002.0, status="CLOSED",
                    exit_reason=exit_reason, exit_ts_utc=exit_ts,
                    pnl_pips=pnl, pnl_usd=round(pnl * 2, 2),
                    mae_pips=mae, mfe_pips=mfe,
                    news_flag=(trade_num % 3 == 0),   # ~1/3 with news
                    vol_regime="normal",
                    spread_at_entry_pips=1.5,
                )
                store.insert_trade(trade)
                trade_num += 1

        # Add some skipped signals for the funnel
        for status, n in [("SKIPPED_SPREAD", 2), ("SKIPPED_SL_TOO_WIDE", 1),
                           ("SKIPPED_POSITION_OPEN", 1)]:
            for _ in range(n):
                skip_sig = Signal(
                    signal_id=uuid.uuid4().hex,
                    ts_utc=base_entry + timedelta(hours=trade_num + 100),
                    strategy=strategy, direction=direction, killzone=killzone,
                    entry_type="MARKET", entry_price=2000.0,
                    entry_zone_low=1999.5, entry_zone_high=2000.5,
                    sl=1998.0, tp1=2001.0, tp2=2002.0,
                    sl_pips=20.0, confluences=[], score=2, context={},
                )
                store.insert_signal(skip_sig, status=status, skip_reason="test")
                trade_num += 1

    yield


# ── Metric value tests ────────────────────────────────────────────────────────

class TestComputeStats:

    def test_total_count(self, seeded):
        assert compute_stats().count == 40

    def test_per_strategy_count(self, seeded):
        for s in ["S1", "S2", "S3", "S4"]:
            assert compute_stats(strategy=s).count == 10

    def test_winrate(self, seeded):
        st = compute_stats(strategy="S1")
        assert st.wins == 6
        assert st.losses == 4
        assert st.winrate == pytest.approx(0.6, abs=0.001)

    def test_avg_win_pips(self, seeded):
        st = compute_stats(strategy="S1")
        # (4×18 + 2×0.5) / 6 = 73/6
        assert st.avg_win_pips == pytest.approx(73/6, abs=0.01)

    def test_avg_loss_pips(self, seeded):
        st = compute_stats(strategy="S1")
        # (3×(-15) + 1×(-3)) / 4 = -12.0
        assert st.avg_loss_pips == pytest.approx(-12.0, abs=0.01)

    def test_expectancy(self, seeded):
        st = compute_stats(strategy="S1")
        # 0.6×12.167 + 0.4×(-12) = 2.5
        assert st.expectancy_pips == pytest.approx(2.5, abs=0.1)

    def test_profit_factor(self, seeded):
        st = compute_stats(strategy="S1")
        # 73 / 48 ≈ 1.521
        assert st.profit_factor == pytest.approx(73/48, abs=0.01)

    def test_total_pnl_all_strategies(self, seeded):
        st = compute_stats()
        # Per strategy: 4×18 + 2×0.5 + 3×(-15) + 1×(-3) = 72+1-45-3 = 25 pips
        # 4 strategies × 25 = 100 pips
        assert st.total_pnl_pips == pytest.approx(100.0, abs=0.5)

    def test_exit_reason_breakdown(self, seeded):
        st = compute_stats(strategy="S1")
        assert st.exits["TP2"] == 4
        assert st.exits["BE"] == 2
        assert st.exits["SL"] == 3
        assert st.exits["TIMEOUT"] == 1
        assert st.exits["MANUAL"] == 0

    def test_exit_pct_sums_to_one(self, seeded):
        st = compute_stats(strategy="S1")
        total = sum(st.exits_pct.values())
        assert total == pytest.approx(1.0, abs=0.001)

    def test_mae_mfe_stats(self, seeded):
        st = compute_stats(strategy="S1")
        # Winners (6): 4×mae=3 + 2×mae=2 → avg = (12+4)/6 = 16/6 ≈ 2.667
        assert st.avg_mae_winners == pytest.approx(16/6, abs=0.05)
        # Losers (4): 3×mae=15 + 1×mae=5 → avg = (45+5)/4 = 12.5
        assert st.avg_mae_losers == pytest.approx(12.5, abs=0.05)
        # Losers MFE: 3×mfe=5 + 1×mfe=8 → avg = (15+8)/4 = 5.75
        assert st.avg_mfe_losers == pytest.approx(5.75, abs=0.05)

    def test_avg_duration_minutes(self, seeded):
        st = compute_stats(strategy="S1")
        # 4×28 + 2×22 + 3×42 + 1×46 = 112+44+126+46 = 328 → avg = 32.8
        assert st.avg_duration_min == pytest.approx(32.8, abs=0.5)

    def test_empty_result_for_no_trades(self, seeded):
        st = compute_stats(strategy="S1", direction="SHORT")
        # S1 is LONG-only in fixture
        assert st.count == 0
        assert st.winrate == 0.0
        assert st.profit_factor == 0.0

    def test_direction_filter(self, seeded):
        # S4 is SHORT
        st = compute_stats(strategy="S4", direction="SHORT")
        assert st.count == 10

    def test_killzone_filter(self, seeded):
        # S1 and S4 are both LONDON
        st = compute_stats(killzone="LONDON")
        assert st.count == 20   # 10 S1 + 10 S4

    def test_news_flag_filter(self, seeded):
        # news_flag set to True for trade_num % 3 == 0 → roughly 4 per strategy
        st_no  = compute_stats(strategy="S1", news_flag=0)
        st_yes = compute_stats(strategy="S1", news_flag=1)
        assert st_no.count + st_yes.count == 10


# ── Funnel tests ──────────────────────────────────────────────────────────────

class TestComputeFunnel:

    def test_executed_count(self, seeded):
        funnels = {f.strategy: f for f in compute_funnel()}
        for s in ["S1", "S2", "S3", "S4"]:
            assert funnels[s].executed == 10

    def test_skip_counts(self, seeded):
        funnels = {f.strategy: f for f in compute_funnel()}
        for s in ["S1", "S2", "S3", "S4"]:
            f = funnels[s]
            assert f.skipped_spread == 2
            assert f.skipped_sl_too_wide == 1
            assert f.skipped_position_open == 1

    def test_total_signals(self, seeded):
        funnels = {f.strategy: f for f in compute_funnel()}
        for s in ["S1", "S2", "S3", "S4"]:
            # 10 executed + 2 spread + 1 sl + 1 pos = 14
            assert funnels[s].total_signals == 14

    def test_single_strategy_filter(self, seeded):
        result = compute_funnel(strategy="S2")
        assert len(result) == 1
        assert result[0].strategy == "S2"


# ── By-killzone tests ─────────────────────────────────────────────────────────

class TestByKillzone:

    def test_london_s1_count(self, seeded):
        matrix = compute_by_killzone()
        assert matrix[("S1", "LONDON")].count == 10

    def test_ny_am_s2_count(self, seeded):
        matrix = compute_by_killzone()
        assert matrix[("S2", "NY_AM")].count == 10

    def test_no_s1_in_asia(self, seeded):
        matrix = compute_by_killzone()
        assert matrix[("S1", "ASIA")].count == 0

    def test_matrix_covers_all_cells(self, seeded):
        matrix = compute_by_killzone()
        assert len(matrix) == 16  # 4 strategies × 4 killzones


# ── Daily pnl tests ───────────────────────────────────────────────────────────

class TestComputeDaily:

    def test_daily_rows_non_empty(self, seeded):
        rows = compute_daily()
        assert len(rows) > 0

    def test_cumulative_pnl_last_row(self, seeded):
        rows = compute_daily()
        st = compute_stats()
        # Last row's cum_pnl should equal total
        assert rows[-1].cum_pnl_pips == pytest.approx(st.total_pnl_pips, abs=0.5)

    def test_max_intraday_dd_non_negative(self, seeded):
        for row in compute_daily():
            assert row.max_intraday_dd_pips >= 0.0

    def test_total_trades_in_daily_equals_total(self, seeded):
        rows = compute_daily()
        assert sum(r.trades for r in rows) == 40


# ── Excel export smoke test ───────────────────────────────────────────────────

class TestExcelExport:

    def test_export_creates_file_with_6_sheets(self, seeded, tmp_path):
        from reporting.excel_export import run_export
        from openpyxl import load_workbook

        out = tmp_path / "test_export.xlsx"
        run_export(str(out))

        assert out.exists()
        wb = load_workbook(out)
        assert set(wb.sheetnames) == {
            "Summary", "Trades", "By Killzone",
            "News Split", "Funnel", "Daily"
        }

    def test_summary_sheet_has_data(self, seeded, tmp_path):
        from reporting.excel_export import run_export
        from openpyxl import load_workbook

        out = tmp_path / "test2.xlsx"
        run_export(str(out))
        wb = load_workbook(out)
        ws = wb["Summary"]
        # Row 1 = headers; row 2 = S1 data; row 5 = TOTAL
        assert ws.cell(2, 1).value == "S1"
        assert ws.cell(6, 1).value == "TOTAL"
        # Count for S1 should be 10
        assert ws.cell(2, 2).value == 10

    def test_trades_sheet_has_40_rows(self, seeded, tmp_path):
        from reporting.excel_export import run_export
        from openpyxl import load_workbook

        out = tmp_path / "test3.xlsx"
        run_export(str(out))
        wb = load_workbook(out)
        ws = wb["Trades"]
        # Row 1 = headers, rows 2..41 = trades
        data_rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]
        assert len(data_rows) == 40


# ── Dashboard smoke test ──────────────────────────────────────────────────────

class TestDashboard:

    def test_index_returns_html(self, seeded):
        from fastapi.testclient import TestClient
        from reporting.dashboard import app

        client = TestClient(app)
        resp = client.get("/")
        assert resp.status_code == 200
        assert "xauusd-scalper" in resp.text
        assert "<table>" in resp.text

    def test_api_summary_json(self, seeded):
        from fastapi.testclient import TestClient
        from reporting.dashboard import app

        client = TestClient(app)
        resp = client.get("/api/summary")
        assert resp.status_code == 200
        data = resp.json()
        assert set(data.keys()) >= {"S1", "S2", "S3", "S4"}
        assert data["S1"]["count"] == 10

    def test_api_trades_json(self, seeded):
        from fastapi.testclient import TestClient
        from reporting.dashboard import app

        client = TestClient(app)
        resp = client.get("/api/trades?limit=50")
        assert resp.status_code == 200
        trades = resp.json()
        assert len(trades) == 40
