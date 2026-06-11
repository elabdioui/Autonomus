"""Excel journal export — 6 sheets, regenerated daily.

Run standalone:  python -m reporting.excel_export
Scheduled:       called from main.py at 22:30 UTC via APScheduler.
"""
from __future__ import annotations

import logging
import sqlite3
from datetime import datetime, timezone
from pathlib import Path

import pytz
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

import core.store as _store
from config import cfg
from reporting.metrics import (
    TradeStats, _STRATEGIES, _KILLZONES, _EXIT_REASONS,
    compute_stats, compute_funnel, compute_by_killzone, compute_daily,
)

log = logging.getLogger(__name__)

_LOCAL_TZ   = pytz.timezone(cfg.TIMEZONE)
_GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
_RED_FILL   = PatternFill("solid", fgColor="FFC7CE")
_HEADER_FONT = Font(bold=True)


# ── Formatting helpers ────────────────────────────────────────────────────────

def _h(ws, row: int, cols: list[str]) -> None:
    """Write bold header row."""
    for c, val in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = _HEADER_FONT


def _pct(v: float) -> str:
    return f"{v*100:.1f}%"


def _fmt_pf(v: float) -> str:
    return "∞" if v == float("inf") else f"{v:.2f}"


def _stats_row(s: TradeStats) -> list:
    return [
        s.count, s.wins, s.losses, _pct(s.winrate),
        round(s.avg_win_pips, 2), round(s.avg_loss_pips, 2),
        round(s.expectancy_pips, 2), _fmt_pf(s.profit_factor),
        round(s.total_pnl_pips, 1), round(s.total_pnl_usd, 2),
        s.exits.get("TP2", 0), s.exits.get("BE", 0),
        s.exits.get("SL", 0), s.exits.get("TIMEOUT", 0),
        round(s.avg_duration_min, 1),
        round(s.avg_mae_winners, 2), round(s.avg_mae_losers, 2),
        round(s.avg_mfe_losers, 2),
    ]


def _local_ts(iso: str | None) -> str:
    if not iso:
        return ""
    try:
        dt = datetime.fromisoformat(iso)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(_LOCAL_TZ).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return iso or ""


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 30)


# ── Sheet builders ────────────────────────────────────────────────────────────

_SUMMARY_COLS = [
    "Strategy", "Count", "Wins", "Losses", "Winrate%",
    "AvgWin", "AvgLoss", "Expectancy", "ProfitFactor",
    "TotalPipsPnl", "TotalUsdPnl",
    "TP2#", "BE#", "SL#", "Timeout#",
    "AvgDurMin", "AvgMAE_Win", "AvgMAE_Loss", "AvgMFE_Loss",
]


def _sheet_summary(wb: Workbook) -> None:
    ws = wb.create_sheet("Summary")
    _h(ws, 1, _SUMMARY_COLS)

    exp_col = _SUMMARY_COLS.index("Expectancy") + 1   # 1-based

    row = 2
    totals: list[TradeStats] = []
    for s in _STRATEGIES:
        st = compute_stats(strategy=s)
        totals.append(st)
        ws.append([s] + _stats_row(st))
        # Conditional format on the Expectancy cell of this row
        cell = ws.cell(row=row, column=exp_col)
        cell.fill = _GREEN_FILL if st.expectancy_pips > 0 else _RED_FILL
        row += 1

    # TOTAL row
    all_st = compute_stats()
    ws.append(["TOTAL"] + _stats_row(all_st))
    cell = ws.cell(row=row, column=exp_col)
    cell.fill = _GREEN_FILL if all_st.expectancy_pips > 0 else _RED_FILL

    _auto_width(ws)


def _sheet_trades(wb: Workbook) -> None:
    ws = wb.create_sheet("Trades")
    cols = [
        "TradeId", "Strategy", "Direction", "Killzone", "EntryType",
        "Lot", "EntryFill", "Entry(local)", "ExitTs(local)",
        "SlInitial", "SlCurrent", "Tp1", "Tp2",
        "Status", "ExitReason",
        "PnlPips", "PnlUsd", "MaePips", "MfePips",
        "NewsFlag", "VolRegime", "SpreadEntry",
    ]
    _h(ws, 1, cols)

    con = sqlite3.connect(_store.DB_PATH)
    con.row_factory = sqlite3.Row
    rows = con.execute(
        """SELECT t.*, s.killzone, s.entry_type
           FROM trades t
           LEFT JOIN signals s ON t.signal_id = s.signal_id
           ORDER BY t.entry_ts_utc DESC"""
    ).fetchall()
    con.close()

    for r in rows:
        ws.append([
            r["trade_id"][:8],
            r["strategy"], r["direction"],
            r["killzone"] or "", r["entry_type"] or "",
            r["lot"], r["entry_price_fill"],
            _local_ts(r["entry_ts_utc"]), _local_ts(r["exit_ts_utc"]),
            r["sl_initial"], r["sl_current"], r["tp1"], r["tp2"],
            r["status"], r["exit_reason"],
            r["pnl_pips"], r["pnl_usd"],
            r["mae_pips"], r["mfe_pips"],
            bool(r["news_flag"]), r["vol_regime"], r["spread_at_entry_pips"],
        ])

    _auto_width(ws)


def _sheet_by_killzone(wb: Workbook) -> None:
    ws = wb.create_sheet("By Killzone")
    header = ["Strategy \\ Killzone"]
    for kz in _KILLZONES:
        header += [f"{kz} N", f"{kz} WR%", f"{kz} Exp"]
    _h(ws, 1, header)

    matrix = compute_by_killzone()
    for s in _STRATEGIES:
        row = [s]
        for kz in _KILLZONES:
            st = matrix[(s, kz)]
            row += [st.count, _pct(st.winrate), round(st.expectancy_pips, 2)]
        ws.append(row)

    _auto_width(ws)


def _sheet_news_split(wb: Workbook) -> None:
    ws = wb.create_sheet("News Split")
    header = ["Strategy"]
    for label in ("NoNews_", "RedNews_"):
        header += [f"{label}N", f"{label}WR%", f"{label}Exp"]
    _h(ws, 1, header)

    for s in _STRATEGIES:
        no_news = compute_stats(strategy=s, news_flag=0)
        red     = compute_stats(strategy=s, news_flag=1)
        ws.append([
            s,
            no_news.count, _pct(no_news.winrate), round(no_news.expectancy_pips, 2),
            red.count,     _pct(red.winrate),      round(red.expectancy_pips, 2),
        ])

    _auto_width(ws)


def _sheet_funnel(wb: Workbook) -> None:
    ws = wb.create_sheet("Funnel")
    cols = [
        "Strategy", "TotalSignals", "Detected", "Executed",
        "Skip_SL", "Skip_PosOpen", "Skip_Cooldown", "Skip_Spread", "Skip_Rejected",
    ]
    _h(ws, 1, cols)

    for f in compute_funnel():
        ws.append([
            f.strategy, f.total_signals, f.detected, f.executed,
            f.skipped_sl_too_wide, f.skipped_position_open,
            f.skipped_cooldown, f.skipped_spread, f.skipped_order_rejected,
        ])

    _auto_width(ws)


def _sheet_daily(wb: Workbook) -> None:
    ws = wb.create_sheet("Daily")
    cols = ["Date", "Trades", "PnlPips", "PnlUsd", "CumPnlPips", "MaxIntradayDD"]
    _h(ws, 1, cols)

    for d in compute_daily(cfg.TIMEZONE):
        ws.append([
            d.date, d.trades,
            d.pnl_pips, d.pnl_usd,
            d.cum_pnl_pips, d.max_intraday_dd_pips,
        ])

    _auto_width(ws)


# ── Public entry point ────────────────────────────────────────────────────────

def run_export(path: str | None = None) -> Path:
    out = Path(path or cfg.EXCEL_EXPORT_PATH)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)    # remove default empty sheet

    _sheet_summary(wb)
    _sheet_trades(wb)
    _sheet_by_killzone(wb)
    _sheet_news_split(wb)
    _sheet_funnel(wb)
    _sheet_daily(wb)

    wb.save(out)
    log.info("Excel export saved → %s", out)
    return out


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    run_export()
